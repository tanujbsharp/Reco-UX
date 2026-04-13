"""
Excel upload handler for bulk product import.

Accepts an .xlsx file, parses columns, and creates/updates Product
and FeatureValue records for a given Packet.

Expected columns:
  product_code | model | family | price | product_url | <feature_code_1> | <feature_code_2> | ...

Products whose product_url is non-empty are automatically marked
crawl_status='pending' so the crawl pipeline can pick them up.
"""
import logging

from openpyxl import load_workbook

from apps.packets.models import Feature, FeatureValue, Product

logger = logging.getLogger(__name__)

# Columns that map directly to Product model fields
PRODUCT_FIELD_COLUMNS = {'product_code', 'model', 'family', 'price', 'product_url'}


def import_products_from_excel(file_obj, packet):
    """
    Parse an .xlsx file and create/update Products + FeatureValues.

    Args:
        file_obj: An in-memory uploaded file (InMemoryUploadedFile or similar).
        packet: The Packet instance to attach products to.

    Returns:
        dict with 'created', 'updated', 'errors' counts.
    """
    wb = load_workbook(filename=file_obj, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'created': 0, 'updated': 0, 'errors': 0, 'detail': 'Empty spreadsheet'}

    headers = [str(h).strip().lower() if h else '' for h in rows[0]]

    # Validate required columns
    missing = PRODUCT_FIELD_COLUMNS - {'product_url'} - set(headers)
    if missing:
        return {
            'created': 0,
            'updated': 0,
            'errors': 1,
            'detail': f'Missing required columns: {", ".join(sorted(missing))}',
        }

    # Build a lookup of feature_code -> Feature for this packet
    feature_codes_in_sheet = set(headers) - PRODUCT_FIELD_COLUMNS - {''}
    features_by_code = {
        f.feature_code.lower(): f
        for f in Feature.objects.filter(packet=packet, feature_code__in=[
            c for c in feature_codes_in_sheet
        ])
    }

    created = 0
    updated = 0
    errors = 0

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None for cell in row):
            continue

        row_dict = {}
        for idx, header in enumerate(headers):
            if header and idx < len(row):
                row_dict[header] = row[idx]

        product_code = row_dict.get('product_code')
        if not product_code:
            logger.warning('Row %d: missing product_code, skipping', row_num)
            errors += 1
            continue

        product_code = str(product_code).strip()

        # Prepare product fields
        product_url = str(row_dict.get('product_url', '') or '').strip()
        crawl_status = 'pending' if product_url else 'not_applicable'

        product_data = {
            'packet': packet,
            'model': str(row_dict.get('model', '')).strip(),
            'family': str(row_dict.get('family', '')).strip(),
            'price': row_dict.get('price', 0) or 0,
            'product_url': product_url,
            'crawl_status': crawl_status,
        }

        try:
            product, was_created = Product.objects.update_or_create(
                product_code=product_code,
                defaults=product_data,
            )
            if was_created:
                created += 1
            else:
                updated += 1
        except Exception as exc:
            logger.error('Row %d: failed to save product %s — %s', row_num, product_code, exc)
            errors += 1
            continue

        # Create/update FeatureValues for any feature columns in the sheet
        for col_name, feature in features_by_code.items():
            raw_value = row_dict.get(col_name)
            if raw_value is None:
                continue

            value_str = str(raw_value).strip()
            normalized = None
            try:
                normalized = float(raw_value)
            except (ValueError, TypeError):
                pass

            FeatureValue.objects.update_or_create(
                product=product,
                feature=feature,
                defaults={
                    'value': value_str,
                    'normalized_value': normalized,
                },
            )

    wb.close()

    return {
        'created': created,
        'updated': updated,
        'errors': errors,
    }
